# -----------------------------------------------------------------------------
# Project Programacion para mortales
# Desarrollador : Kevin Turkienich
# 2024
# -----------------------------------------------------------------------------
# Vista para importación / exportación masiva de productos via Excel

import io
import os
import tempfile
import time
import uuid
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from django.contrib import admin, messages
from django.contrib.admin.views.decorators import staff_member_required
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render, redirect

from .models import (
    Producto,
    Categoria,
    UNIDADES_DE_MEDIDA,
    Receta,
    CategoriaReceta,
    ProductoReceta,
    Bien,
    BienReceta,
)
from .services_bienes import TIME_UNIT_CHOICES, obtener_precio_kwh
from configuracion.models import Configuracion


# ---------------------------------------------------------------------------
# Columnas del Excel (orden fijo)
# La columna ID va primero — solo se rellena en exportación, no en plantilla vacía.
COLUMNAS = [
    ('id',              'ID'),
    ('codigo',          'Código'),
    ('nombre',          'Nombre *'),
    ('descripcion',     'Descripción'),
    ('categoria',       'Categoría'),
    ('marca',           'Marca'),
    ('unidad_de_medida','Unidad de Medida *'),
    ('cantidad',        'Cantidad *'),
    ('costo',           'Costo *'),
]

UNIDADES_VALIDAS = [u[0] for u in UNIDADES_DE_MEDIDA]
UNIDADES_TIEMPO_VALIDAS = [u[0] for u in TIME_UNIT_CHOICES]
OPCIONES_BOOLEANAS_IMPORTACION = ['Si', 'No']


def _apply_header_style(cell):
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill(fill_type='solid', fgColor='2C3E50')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='FFFFFF')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _build_workbook(productos=None, solo_plantilla=False):
    """
    Construye un workbook de openpyxl.
    - `productos=None` + `solo_plantilla=True` → plantilla vacía con dropdown.
    - `productos=[...]`                          → exportación de datos reales.
    """
    wb = openpyxl.Workbook()

    # ---- Hoja oculta con lista de unidades (para el dropdown) ----
    ws_listas = wb.create_sheet(title='_Listas')
    ws_listas.sheet_state = 'hidden'
    for i, unidad in enumerate(UNIDADES_VALIDAS, start=1):
        ws_listas.cell(row=i, column=1, value=unidad)

    # ---- Hoja principal ----
    ws = wb.active
    ws.title = 'Productos'

    # Encabezados
    for col_idx, (_, header) in enumerate(COLUMNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _apply_header_style(cell)

    # Anchos de columna (ID, Código, Nombre, Descripción, Categoría, Marca, Unidad, Cantidad, Costo)
    anchos = [8, 12, 30, 35, 20, 20, 22, 12, 12]
    for col_idx, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'

    # ---- Validación desplegable para Unidad de Medida (columna G = índice 7 ahora) ----
    col_letra_udm = get_column_letter(7)
    rango_listas = f"'_Listas'!$A$1:$A${len(UNIDADES_VALIDAS)}"
    dv = DataValidation(
        type='list',
        formula1=rango_listas,
        allow_blank=False,
        showErrorMessage=True,
        errorTitle='Unidad inválida',
        error='Seleccione una unidad de medida de la lista desplegable.',
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.sqref = f'{col_letra_udm}2:{col_letra_udm}10000'

    if solo_plantilla:
        # Fila de ejemplo (en gris claro): ID vacío (producto nuevo), luego datos
        ejemplo = ['', 'PROD-001', 'Harina 000', 'Harina común tipo 000', 'Almacén', 'Molinos', 'Kilos', '25', '800']
        fill_ejemplo = PatternFill(fill_type='solid', fgColor='F0F0F0')
        italic_font = Font(italic=True, color='888888')
        for col_idx, valor in enumerate(ejemplo, start=1):
            cell = ws.cell(row=2, column=col_idx, value=valor)
            cell.fill = fill_ejemplo
            cell.font = italic_font
    else:
        # Datos reales — ID bloqueado (solo lectura visual, color gris)
        fill_par = PatternFill(fill_type='solid', fgColor='F7F9FC')
        fill_id  = PatternFill(fill_type='solid', fgColor='E8E8E8')
        font_id  = Font(bold=True, color='555555')
        for row_idx, prod in enumerate(productos, start=2):
            fill = fill_par if row_idx % 2 == 0 else None
            valores = [
                prod.pk,
                prod.codigo or '',
                prod.nombre or '',
                prod.descripcion or '',
                prod.categoria.nombre if prod.categoria else '',
                prod.marca or '',
                prod.unidad_de_medida or '',
                float(prod.cantidad) if prod.cantidad is not None else 1,
                float(prod.costo) if prod.costo is not None else 0,
            ]
            for col_idx, valor in enumerate(valores, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                if col_idx == 1:  # columna ID
                    cell.fill = fill_id
                    cell.font = font_id
                elif fill:
                    cell.fill = fill

    return wb


@staff_member_required
def importar_exportar_productos(request):
    """Vista principal para importación / exportación masiva de productos."""

    usuario = request.user.username
    config = Configuracion.objects.filter(usuario=usuario).first() or Configuracion.objects.first()

    # Superusuario ve todos los productos; usuario normal solo los suyos
    if request.user.is_superuser:
        productos_qs = Producto.objects.all()
    else:
        productos_qs = Producto.objects.filter(usuario=usuario)

    if request.method == 'POST':
        accion = request.POST.get('accion', '')

        # ------------------------------------------------------------------
        # EXPORTAR PLANTILLA (vacía)
        # ------------------------------------------------------------------
        if accion == 'exportar_plantilla':
            wb = _build_workbook(solo_plantilla=True)
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="plantilla_productos.xlsx"'
            wb.save(response)
            return response

        # ------------------------------------------------------------------
        # EXPORTAR PRODUCTOS REALES
        # ------------------------------------------------------------------
        if accion == 'exportar_productos':
            productos = productos_qs.order_by('nombre')
            wb = _build_workbook(productos=productos)
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="productos.xlsx"'
            wb.save(response)
            return response

        # ------------------------------------------------------------------
        # IMPORTAR
        # ------------------------------------------------------------------
        if accion == 'importar':
            archivo = request.FILES.get('archivo_excel')
            if not archivo:
                messages.error(request, 'Debe seleccionar un archivo Excel.')
                return redirect('importar_exportar_productos')

            # Validar extensión
            nombre_archivo = archivo.name.lower()
            if not nombre_archivo.endswith('.xlsx'):
                messages.error(request, 'Solo se aceptan archivos .xlsx')
                return redirect('importar_exportar_productos')

            try:
                wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
                ws = wb.active
            except Exception as e:
                messages.error(request, f'No se pudo leer el archivo: {e}')
                return redirect('importar_exportar_productos')

            errores = []
            creados = 0
            actualizados = 0

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Ignorar filas completamente vacías
                if not any(row):
                    continue

                # Columna 0 = ID, 1 = Código, 2 = Nombre ... (nueva estructura)
                id_raw      = row[0]
                codigo      = str(row[1]).strip() if row[1] is not None else ''
                nombre      = str(row[2]).strip() if row[2] is not None else ''
                descripcion = str(row[3]).strip() if row[3] is not None else ''
                cat_nombre  = str(row[4]).strip() if row[4] is not None else ''
                marca       = str(row[5]).strip() if row[5] is not None else ''
                unidad      = str(row[6]).strip() if row[6] is not None else ''
                cantidad_raw = row[7]
                costo_raw    = row[8]

                # Convertir ID a entero si viene del Excel
                producto_id = None
                if id_raw is not None and str(id_raw).strip() not in ('', 'ID'):
                    try:
                        producto_id = int(float(str(id_raw).strip()))
                    except (ValueError, TypeError):
                        producto_id = None

                # ---- Validaciones obligatorias ----
                fila_errores = []

                if not nombre:
                    fila_errores.append('Nombre es obligatorio')

                if unidad not in UNIDADES_VALIDAS:
                    fila_errores.append(
                        f'Unidad de medida "{unidad}" no válida. '
                        f'Usar: {", ".join(UNIDADES_VALIDAS)}'
                    )

                try:
                    cantidad = Decimal(str(cantidad_raw)) if cantidad_raw is not None else Decimal('1')
                    if cantidad <= 0:
                        raise ValueError
                except (InvalidOperation, ValueError):
                    fila_errores.append('Cantidad debe ser un número positivo')
                    cantidad = Decimal('1')

                try:
                    costo = float(costo_raw) if costo_raw is not None else 0.0
                    if costo < 0:
                        raise ValueError
                except (TypeError, ValueError):
                    fila_errores.append('Costo debe ser un número mayor o igual a 0')
                    costo = 0.0

                if fila_errores:
                    errores.append(f'Fila {row_num}: ' + ' | '.join(fila_errores))
                    continue

                # ---- Resolver categoría (crear si no existe) ----
                categoria = None
                if cat_nombre:
                    categoria, _ = Categoria.objects.get_or_create(
                        nombre=cat_nombre,
                        usuario=usuario,
                    )

                # ---- Crear o actualizar producto ----
                # Prioridad: 1) ID de base de datos, 2) Código, 3) Nombre
                producto_qs = None

                if producto_id:
                    producto_qs = Producto.objects.filter(pk=producto_id).first()

                if producto_qs is None and codigo:
                    q = Producto.objects.filter(codigo=codigo)
                    if not request.user.is_superuser:
                        q = q.filter(usuario=usuario)
                    producto_qs = q.first()

                if producto_qs is None:
                    q = Producto.objects.filter(nombre=nombre)
                    if not request.user.is_superuser:
                        q = q.filter(usuario=usuario)
                    producto_qs = q.first()

                if producto_qs:
                    # Actualizar
                    producto_qs.codigo      = codigo or producto_qs.codigo
                    producto_qs.descripcion = descripcion
                    producto_qs.categoria   = categoria
                    producto_qs.marca       = marca
                    producto_qs.unidad_de_medida = unidad
                    producto_qs.cantidad    = cantidad
                    producto_qs.costo       = costo
                    producto_qs.save()
                    actualizados += 1
                else:
                    # Crear nuevo
                    Producto.objects.create(
                        codigo=codigo or None,
                        nombre=nombre,
                        descripcion=descripcion,
                        categoria=categoria,
                        marca=marca,
                        unidad_de_medida=unidad,
                        cantidad=cantidad,
                        costo=costo,
                        usuario=usuario,
                    )
                    creados += 1

            # ---- Resumen ----
            if errores:
                for err in errores[:20]:   # limitar a 20 errores visibles
                    messages.warning(request, err)
                if len(errores) > 20:
                    messages.warning(request, f'… y {len(errores) - 20} errores más no mostrados.')

            if creados or actualizados:
                messages.success(
                    request,
                    f'Importación completada: {creados} producto(s) nuevo(s), '
                    f'{actualizados} actualizado(s).'
                )
            elif not errores:
                messages.info(request, 'No se encontraron filas con datos para importar.')

            return redirect('importar_exportar_productos')

    # GET
    total_productos = productos_qs.count()
    context = {
        'title': 'Importar / Exportar Productos',
        'available_apps': admin.site.get_app_list(request),
        'total_productos': total_productos,
        'unidades_validas': UNIDADES_VALIDAS,
        'config': config,
    }
    return render(request, 'admin/importar_exportar_productos.html', context)


# ===========================================================================
# IMPORTAR / EXPORTAR RECETAS
# ===========================================================================

COLUMNAS_RECETA_IMPORTACION = [
    ('id', 'ID'),
    ('nombre', 'Nombre *'),
    ('descripcion', 'Descripción *'),
    ('categoria', 'Categoría'),
    ('porciones', 'Porciones *'),
    ('rentabilidad', 'Rentabilidad % *'),
    ('iva', 'IVA % *'),
    ('comentarios', 'Comentarios'),
]

COLUMNAS_RECETA_PRODUCTO = [
    ('id', 'ID'),
    ('receta_id', 'Receta ID'),
    ('receta_nombre', 'Receta Nombre'),
    ('producto_id', 'Producto ID'),
    ('producto_nombre', 'Producto Nombre'),
    ('cantidad', 'Cantidad *'),
    ('medida_uso', 'Medida de uso *'),
]

COLUMNAS_BIEN_IMPORTACION = [
    ('id', 'ID'),
    ('nombre', 'Nombre *'),
    ('descripcion', 'Descripción'),
    ('costo_compra', 'Costo de compra *'),
    ('vida_util_cantidad', 'Vida útil cantidad *'),
    ('vida_util_unidad', 'Vida útil unidad *'),
    ('potencia_watts', 'Potencia watts'),
    ('factor_uso_porcentaje', 'Factor de uso %'),
    ('activo', 'Activo'),
]

COLUMNAS_RECETA_BIEN = [
    ('id', 'ID'),
    ('receta_id', 'Receta ID'),
    ('receta_nombre', 'Receta Nombre'),
    ('bien_id', 'Bien ID'),
    ('bien_nombre', 'Bien Nombre'),
    ('tiempo_uso_cantidad', 'Tiempo de uso cantidad *'),
    ('tiempo_uso_unidad', 'Tiempo de uso unidad *'),
    ('incluir_depreciacion', 'Incluir depreciación'),
    ('incluir_electricidad', 'Incluir electricidad'),
    ('observaciones', 'Observaciones'),
]

COLUMNAS_CONFIGURACION_RECETAS = [
    ('precio_kwh', 'Precio kWh *'),
]

COLUMNAS_PRODUCTOS_REFERENCIA = [
    ('id', 'ID'),
    ('nombre', 'Nombre'),
    ('unidad_de_medida', 'Unidad de medida'),
    ('categoria', 'Categoría'),
    ('marca', 'Marca'),
    ('cantidad', 'Cantidad'),
    ('costo', 'Costo'),
]

HOJA_RECETAS = 'Recetas'
HOJA_RECETAS_PRODUCTOS = 'Recetas-Productos'
HOJA_BIENES = 'Bienes'
HOJA_RECETAS_BIENES = 'Recetas-Bienes'
HOJA_CONFIGURACION_RECETAS = 'Configuracion'
HOJA_PRODUCTOS = 'Productos'
HOJA_LISTAS_RECETAS = '_ListasRecetas'

PREVIEW_RECETAS_SESSION_KEY = 'preview_importacion_recetas'
PREVIEW_RECETAS_MAX_ERRORES = 50
PREVIEW_RECETAS_MAX_CAMBIOS = 25
PREVIEW_RECETAS_WARN_UMBRAL = 500
PREVIEW_INGREDIENTES_WARN_UMBRAL = 3000
PREVIEW_BIENES_WARN_UMBRAL = 1500

UNIDADES_COMPATIBLES_PRODUCTO_RECETA = {
    'Unidades': {'Unidades'},
    'Mt2s': {'Mt2s'},
    'Kilos': {'Kilos', 'Gramos'},
    'Gramos': {'Kilos', 'Gramos'},
    'Litros': {'Litros', 'Mililitros'},
    'Mililitros': {'Litros', 'Mililitros'},
    'Onzas': {'Onzas', 'Libras'},
    'Libras': {'Onzas', 'Libras'},
    'Metros': {'Metros', 'Centimetros'},
    'Centimetros': {'Metros', 'Centimetros'},
}


def _apply_header_style_receta(cell, readonly=False):
    color = '7F8C8D' if readonly else '2C3E50'
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill(fill_type='solid', fgColor=color)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='FFFFFF')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _normalize_lookup_text(value):
    if value is None:
        return ''
    return str(value).strip().lower()


def _normalize_categoria_receta(value):
    texto = str(value).strip() if value is not None else ''
    return texto.capitalize() if texto else ''


def _row_has_values(row):
    return any(value is not None and str(value).strip() != '' for value in row)


def _parse_excel_int(value):
    if value is None:
        return None
    texto = str(value).strip()
    if not texto:
        return None
    return int(float(texto))


def _parse_excel_decimal(value, *, default=None, min_value=None, max_exclusive=None):
    if value is None or str(value).strip() == '':
        if default is None:
            raise InvalidOperation()
        value = default

    decimal_value = Decimal(str(value))
    if min_value is not None and decimal_value < min_value:
        raise ValueError()
    if max_exclusive is not None and decimal_value >= max_exclusive:
        raise ValueError()
    return decimal_value


def _parse_excel_bool(value, *, default=None):
    if value is None or str(value).strip() == '':
        if default is None:
            raise ValueError()
        return default

    if isinstance(value, bool):
        return value

    texto = str(value).strip().lower()
    if texto in {'si', 'sí', 's', 'true', '1', 'x', 'yes'}:
        return True
    if texto in {'no', 'n', 'false', '0'}:
        return False
    raise ValueError()


def _safe_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def _build_name_lookup(items):
    lookup = {}
    duplicates = set()
    for item in items:
        key = _normalize_lookup_text(getattr(item, 'nombre', ''))
        if not key:
            continue
        if key in lookup:
            duplicates.add(key)
        else:
            lookup[key] = item

    for key in duplicates:
        lookup[key] = None
    return lookup


def _worksheet_headers(ws, expected_columns):
    return [str(ws.cell(row=1, column=index).value or '').strip() for index in range(1, len(expected_columns) + 1)]


def _iter_sheet_rows(ws, expected_columns):
    total_columns = len(expected_columns)
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        normalized = list(row[:total_columns]) if row else []
        if len(normalized) < total_columns:
            normalized.extend([None] * (total_columns - len(normalized)))
        if not _row_has_values(normalized):
            continue
        yield row_num, normalized


def _validate_medida_producto_receta(producto, medida_uso):
    permitidas = UNIDADES_COMPATIBLES_PRODUCTO_RECETA.get(producto.unidad_de_medida, set())
    if medida_uso not in permitidas:
        raise ValueError(
            f"La medida '{medida_uso}' no es compatible con la unidad del producto '{producto.unidad_de_medida}'."
        )


def _preview_action_label(action):
    return 'Actualizar' if action == 'actualizar' else 'Crear'


def _serialize_preview_for_session(preview, temp_path, original_filename):
    return {
        'token': preview['token'],
        'temp_path': temp_path,
        'original_filename': original_filename,
        'summary': preview['summary'],
        'warnings': preview['warnings'],
        'errors': preview['errors'][:PREVIEW_RECETAS_MAX_ERRORES],
        'total_errors': len(preview['errors']),
        'can_confirm': not preview['errors'],
        'configuracion_actions': preview['configuracion_actions'][:PREVIEW_RECETAS_MAX_CAMBIOS],
        'receta_actions': preview['receta_actions'][:PREVIEW_RECETAS_MAX_CAMBIOS],
        'bien_actions': preview['bien_actions'][:PREVIEW_RECETAS_MAX_CAMBIOS],
        'relacion_actions': preview['relacion_actions'][:PREVIEW_RECETAS_MAX_CAMBIOS],
        'bien_relacion_actions': preview['bien_relacion_actions'][:PREVIEW_RECETAS_MAX_CAMBIOS],
        'high_volume': preview['high_volume'],
    }


def _delete_temp_file(path):
    if path and os.path.exists(path):
        try:
            os.remove(path)
        except PermissionError:
            # En Windows el archivo puede quedar bloqueado unos milisegundos más.
            time.sleep(0.1)
            try:
                os.remove(path)
            except PermissionError:
                pass


def _clear_preview_recetas(request, *, remove_file=False):
    preview = request.session.get(PREVIEW_RECETAS_SESSION_KEY)
    if preview and remove_file:
        _delete_temp_file(preview.get('temp_path'))
    if PREVIEW_RECETAS_SESSION_KEY in request.session:
        del request.session[PREVIEW_RECETAS_SESSION_KEY]


def _store_preview_upload(archivo):
    suffix = os.path.splitext(archivo.name)[1] or '.xlsx'
    fd, temp_path = tempfile.mkstemp(suffix=suffix, prefix='recetas_import_')
    try:
        with os.fdopen(fd, 'wb') as temp_file:
            for chunk in archivo.chunks():
                temp_file.write(chunk)
    finally:
        pass
    return temp_path


def _resolve_categoria_receta(nombre, usuario_owner):
    categoria_normalizada = _normalize_categoria_receta(nombre)
    if not categoria_normalizada:
        return None
    categoria = CategoriaReceta.objects.filter(usuario=usuario_owner, nombre__iexact=categoria_normalizada).first()
    if categoria:
        return categoria
    return CategoriaReceta.objects.create(nombre=categoria_normalizada, usuario=usuario_owner)


def _resolve_bien_usuario(request, bien_existente=None):
    if bien_existente and bien_existente.usuario:
        return bien_existente.usuario
    return request.user.username


def _leer_hoja_recetas(ws, recetas_qs, request):
    errores = []
    warnings = []
    planes = []
    acciones = []
    total_filas = 0

    recetas_por_id = {receta.pk: receta for receta in recetas_qs}
    recetas_por_nombre = _build_name_lookup(recetas_qs)
    seen_recipe_keys = {}

    encabezados = _worksheet_headers(ws, COLUMNAS_RECETA_IMPORTACION)
    esperados = [header for _, header in COLUMNAS_RECETA_IMPORTACION]
    if encabezados != esperados:
        errores.append(
            f"Hoja {HOJA_RECETAS}: encabezados inválidos. Se esperaba: {', '.join(esperados)}."
        )
        return {
            'errors': errores,
            'warnings': warnings,
            'plans': planes,
            'actions': acciones,
            'total_rows': total_filas,
            'plans_by_id': {},
            'plans_by_name': {},
        }

    for row_num, row in _iter_sheet_rows(ws, COLUMNAS_RECETA_IMPORTACION):
        total_filas += 1
        id_raw, nombre_raw, descripcion_raw, categoria_raw, porciones_raw, rentabilidad_raw, iva_raw, comentarios_raw = row

        fila_errores = []
        nombre = str(nombre_raw).strip() if nombre_raw is not None else ''
        descripcion = str(descripcion_raw).strip() if descripcion_raw is not None else ''
        comentarios = str(comentarios_raw).strip() if comentarios_raw is not None else ''
        categoria = _normalize_categoria_receta(categoria_raw)
        nombre_lookup = _normalize_lookup_text(nombre)

        receta_id = None
        if id_raw is not None and str(id_raw).strip() != '':
            try:
                receta_id = _parse_excel_int(id_raw)
            except (TypeError, ValueError):
                fila_errores.append('ID de receta inválido')

        if not nombre:
            fila_errores.append('Nombre es obligatorio')

        if not descripcion:
            fila_errores.append('Descripción es obligatoria')

        try:
            porciones = _parse_excel_decimal(
                porciones_raw,
                default=Decimal('1'),
                min_value=Decimal('0.01'),
            )
        except (InvalidOperation, ValueError):
            fila_errores.append('Porciones debe ser un número mayor a 0')
            porciones = Decimal('1')

        try:
            rentabilidad = _parse_excel_decimal(
                rentabilidad_raw,
                default=Decimal('0'),
                min_value=Decimal('0'),
                max_exclusive=Decimal('100'),
            )
        except (InvalidOperation, ValueError):
            fila_errores.append('Rentabilidad debe estar entre 0 y 99.99')
            rentabilidad = Decimal('0')

        try:
            iva = _parse_excel_decimal(
                iva_raw,
                default=Decimal('21'),
                min_value=Decimal('0'),
            )
        except (InvalidOperation, ValueError):
            fila_errores.append('IVA debe ser un número mayor o igual a 0')
            iva = Decimal('21')

        receta_existente = None
        owner_user = request.user.username
        dedupe_key = None

        if receta_id and receta_id not in recetas_por_id:
            fila_errores.append('No existe una receta accesible con ese ID')
        elif receta_id:
            receta_existente = recetas_por_id[receta_id]
            owner_user = receta_existente.usuario or owner_user
            dedupe_key = ('id', receta_existente.pk)
        elif nombre_lookup:
            receta_por_nombre = recetas_por_nombre.get(nombre_lookup)
            if receta_por_nombre is None and nombre_lookup in recetas_por_nombre:
                fila_errores.append('Hay más de una receta existente con ese nombre; use ID para evitar ambigüedad')
            else:
                receta_existente = receta_por_nombre
                if receta_existente:
                    owner_user = receta_existente.usuario or owner_user
                    dedupe_key = ('id', receta_existente.pk)
                else:
                    dedupe_key = ('name', nombre_lookup)

        if categoria and not CategoriaReceta.objects.filter(usuario=owner_user, nombre__iexact=categoria).exists():
            warnings.append(
                f"Hoja {HOJA_RECETAS}, fila {row_num}: se creará la categoría '{categoria}'."
            )

        if dedupe_key in seen_recipe_keys:
            fila_errores.append(
                f"La receta ya fue incluida en la fila {seen_recipe_keys[dedupe_key]} de la hoja {HOJA_RECETAS}"
            )

        if fila_errores:
            errores.append(f"Hoja {HOJA_RECETAS}, fila {row_num}: " + ' | '.join(fila_errores))
            continue

        seen_recipe_keys[dedupe_key] = row_num
        plan_key = f'receta-{row_num}'
        action = 'actualizar' if receta_existente else 'crear'
        plan = {
            'plan_key': plan_key,
            'row_num': row_num,
            'action': action,
            'existing_id': receta_existente.pk if receta_existente else None,
            'owner_user': owner_user,
            'nombre': nombre,
            'nombre_lookup': nombre_lookup,
            'descripcion': descripcion,
            'categoria': categoria,
            'porciones': str(porciones),
            'rentabilidad': str(rentabilidad),
            'iva': str(iva),
            'comentarios': comentarios,
        }
        planes.append(plan)
        acciones.append({
            'sheet': HOJA_RECETAS,
            'row_num': row_num,
            'label': nombre,
            'action': _preview_action_label(action),
        })

    return {
        'errors': errores,
        'warnings': warnings,
        'plans': planes,
        'actions': acciones,
        'total_rows': total_filas,
        'plans_by_id': {plan['existing_id']: plan for plan in planes if plan['existing_id']},
        'plans_by_name': {plan['nombre_lookup']: plan for plan in planes if plan['nombre_lookup']},
    }


def _leer_hoja_configuracion_recetas(ws):
    errores = []
    warnings = []
    actions = []
    total_filas = 0
    plan = None

    encabezados = _worksheet_headers(ws, COLUMNAS_CONFIGURACION_RECETAS)
    esperados = [header for _, header in COLUMNAS_CONFIGURACION_RECETAS]
    if encabezados != esperados:
        errores.append(
            f"Hoja {HOJA_CONFIGURACION_RECETAS}: encabezados inválidos. Se esperaba: {', '.join(esperados)}."
        )
        return {
            'errors': errores,
            'warnings': warnings,
            'actions': actions,
            'total_rows': total_filas,
            'plan': None,
        }

    filas = list(_iter_sheet_rows(ws, COLUMNAS_CONFIGURACION_RECETAS))
    total_filas = len(filas)
    if not filas:
        warnings.append(f'Hoja {HOJA_CONFIGURACION_RECETAS}: no se encontraron datos para precio kWh.')
        return {
            'errors': errores,
            'warnings': warnings,
            'actions': actions,
            'total_rows': total_filas,
            'plan': None,
        }

    if len(filas) > 1:
        errores.append(f'Hoja {HOJA_CONFIGURACION_RECETAS}: solo se permite una fila con datos.')
        return {
            'errors': errores,
            'warnings': warnings,
            'actions': actions,
            'total_rows': total_filas,
            'plan': None,
        }

    row_num, row = filas[0]
    precio_kwh_raw = row[0]
    try:
        precio_kwh = _parse_excel_decimal(
            precio_kwh_raw,
            min_value=Decimal('0'),
        )
    except (InvalidOperation, ValueError):
        errores.append(f'Hoja {HOJA_CONFIGURACION_RECETAS}, fila {row_num}: Precio kWh debe ser un número mayor o igual a 0')
        precio_kwh = Decimal('0')

    if not errores:
        configuracion = Configuracion.objects.first()
        plan = {
            'row_num': row_num,
            'action': 'actualizar' if configuracion else 'crear',
            'existing_id': configuracion.pk if configuracion else None,
            'precio_kwh': str(precio_kwh),
        }
        actions.append({
            'sheet': HOJA_CONFIGURACION_RECETAS,
            'row_num': row_num,
            'label': f'Precio kWh: {precio_kwh}',
            'action': _preview_action_label(plan['action']),
        })

    return {
        'errors': errores,
        'warnings': warnings,
        'actions': actions,
        'total_rows': total_filas,
        'plan': plan,
    }


def _leer_hoja_bienes(ws, bienes_qs, request):
    errores = []
    warnings = []
    planes = []
    acciones = []
    total_filas = 0

    bienes_por_id = {bien.pk: bien for bien in bienes_qs}
    bienes_por_nombre = _build_name_lookup(bienes_qs)
    seen_keys = {}

    encabezados = _worksheet_headers(ws, COLUMNAS_BIEN_IMPORTACION)
    esperados = [header for _, header in COLUMNAS_BIEN_IMPORTACION]
    if encabezados != esperados:
        errores.append(
            f"Hoja {HOJA_BIENES}: encabezados inválidos. Se esperaba: {', '.join(esperados)}."
        )
        return {
            'errors': errores,
            'warnings': warnings,
            'plans': planes,
            'actions': acciones,
            'total_rows': total_filas,
            'plans_by_id': {},
            'plans_by_name': {},
        }

    for row_num, row in _iter_sheet_rows(ws, COLUMNAS_BIEN_IMPORTACION):
        total_filas += 1
        id_raw, nombre_raw, descripcion_raw, costo_compra_raw, vida_util_cantidad_raw, vida_util_unidad_raw, potencia_watts_raw, factor_uso_raw, activo_raw = row

        fila_errores = []
        nombre = str(nombre_raw).strip() if nombre_raw is not None else ''
        descripcion = str(descripcion_raw).strip() if descripcion_raw is not None else ''
        nombre_lookup = _normalize_lookup_text(nombre)

        bien_id = None
        if id_raw is not None and str(id_raw).strip() != '':
            try:
                bien_id = _parse_excel_int(id_raw)
            except (TypeError, ValueError):
                fila_errores.append('ID de bien inválido')

        if not nombre:
            fila_errores.append('Nombre es obligatorio')

        try:
            costo_compra = _parse_excel_decimal(costo_compra_raw, min_value=Decimal('0'))
        except (InvalidOperation, ValueError):
            fila_errores.append('Costo de compra debe ser un número mayor o igual a 0')
            costo_compra = Decimal('0')

        try:
            vida_util_cantidad = _parse_excel_decimal(vida_util_cantidad_raw, min_value=Decimal('0.01'))
        except (InvalidOperation, ValueError):
            fila_errores.append('Vida útil cantidad debe ser un número mayor a 0')
            vida_util_cantidad = Decimal('1')

        vida_util_unidad = str(vida_util_unidad_raw).strip() if vida_util_unidad_raw is not None else ''
        if not vida_util_unidad:
            fila_errores.append('Vida útil unidad es obligatoria')
        elif vida_util_unidad not in UNIDADES_TIEMPO_VALIDAS:
            fila_errores.append('Vida útil unidad inválida')

        try:
            potencia_watts = _parse_excel_decimal(potencia_watts_raw, default=Decimal('0'), min_value=Decimal('0'))
        except (InvalidOperation, ValueError):
            fila_errores.append('Potencia watts debe ser un número mayor o igual a 0')
            potencia_watts = Decimal('0')

        try:
            factor_uso = _parse_excel_decimal(factor_uso_raw, default=Decimal('100'), min_value=Decimal('0'))
            if factor_uso > Decimal('100'):
                raise ValueError()
        except (InvalidOperation, ValueError):
            fila_errores.append('Factor de uso debe estar entre 0 y 100')
            factor_uso = Decimal('100')

        try:
            activo = _parse_excel_bool(activo_raw, default=True)
        except ValueError:
            fila_errores.append("Activo debe ser 'Si' o 'No'")
            activo = True

        bien_existente = None
        dedupe_key = None
        owner_user = request.user.username
        if bien_id and bien_id not in bienes_por_id:
            fila_errores.append('No existe un bien accesible con ese ID')
        elif bien_id:
            bien_existente = bienes_por_id[bien_id]
            owner_user = _resolve_bien_usuario(request, bien_existente)
            dedupe_key = ('id', bien_existente.pk)
        elif nombre_lookup:
            bien_por_nombre = bienes_por_nombre.get(nombre_lookup)
            if bien_por_nombre is None and nombre_lookup in bienes_por_nombre:
                fila_errores.append('Hay más de un bien existente con ese nombre; use ID para evitar ambigüedad')
            else:
                bien_existente = bien_por_nombre
                owner_user = _resolve_bien_usuario(request, bien_existente)
                if bien_existente:
                    dedupe_key = ('id', bien_existente.pk)
                else:
                    dedupe_key = ('name', nombre_lookup)

        if dedupe_key in seen_keys:
            fila_errores.append(
                f"El bien ya fue incluido en la fila {seen_keys[dedupe_key]} de la hoja {HOJA_BIENES}"
            )

        if fila_errores:
            errores.append(f"Hoja {HOJA_BIENES}, fila {row_num}: " + ' | '.join(fila_errores))
            continue

        seen_keys[dedupe_key] = row_num
        plan_key = f'bien-{row_num}'
        action = 'actualizar' if bien_existente else 'crear'
        plan = {
            'plan_key': plan_key,
            'row_num': row_num,
            'action': action,
            'existing_id': bien_existente.pk if bien_existente else None,
            'owner_user': owner_user,
            'nombre': nombre,
            'nombre_lookup': nombre_lookup,
            'descripcion': descripcion,
            'costo_compra': str(costo_compra),
            'vida_util_cantidad': str(vida_util_cantidad),
            'vida_util_unidad': vida_util_unidad,
            'potencia_watts': str(potencia_watts),
            'factor_uso_porcentaje': str(factor_uso),
            'activo': activo,
        }
        planes.append(plan)
        acciones.append({
            'sheet': HOJA_BIENES,
            'row_num': row_num,
            'label': nombre,
            'action': _preview_action_label(action),
        })

    return {
        'errors': errores,
        'warnings': warnings,
        'plans': planes,
        'actions': acciones,
        'total_rows': total_filas,
        'plans_by_id': {plan['existing_id']: plan for plan in planes if plan['existing_id']},
        'plans_by_name': {plan['nombre_lookup']: plan for plan in planes if plan['nombre_lookup']},
    }


def _leer_hoja_recetas_productos(ws, recetas_qs, productos_qs, recetas_preview, request):
    errores = []
    planes = []
    acciones = []
    total_filas = 0

    productos_por_id = {producto.pk: producto for producto in productos_qs}
    productos_por_nombre = _build_name_lookup(productos_qs)
    recetas_por_id = {receta.pk: receta for receta in recetas_qs}
    recetas_por_nombre = _build_name_lookup(recetas_qs)
    relaciones_qs = ProductoReceta.objects.select_related('receta', 'producto')
    if not request.user.is_superuser:
        relaciones_qs = relaciones_qs.filter(receta__usuario=request.user.username)

    relaciones_por_id = {relacion.pk: relacion for relacion in relaciones_qs}
    seen_relation_keys = {}

    encabezados = _worksheet_headers(ws, COLUMNAS_RECETA_PRODUCTO)
    esperados = [header for _, header in COLUMNAS_RECETA_PRODUCTO]
    if encabezados != esperados:
        errores.append(
            f"Hoja {HOJA_RECETAS_PRODUCTOS}: encabezados inválidos. Se esperaba: {', '.join(esperados)}."
        )
        return {
            'errors': errores,
            'plans': planes,
            'actions': acciones,
            'total_rows': total_filas,
        }

    for row_num, row in _iter_sheet_rows(ws, COLUMNAS_RECETA_PRODUCTO):
        total_filas += 1
        id_raw, receta_id_raw, receta_nombre_raw, producto_id_raw, producto_nombre_raw, cantidad_raw, medida_raw = row
        fila_errores = []

        try:
            relacion_id = _parse_excel_int(id_raw) if id_raw is not None and str(id_raw).strip() != '' else None
        except (TypeError, ValueError):
            relacion_id = None
            fila_errores.append('ID de relación inválido')

        receta_plan = None
        receta_obj = None
        receta_lookup = _normalize_lookup_text(receta_nombre_raw)
        if receta_id_raw is not None and str(receta_id_raw).strip() != '':
            try:
                receta_id = _parse_excel_int(receta_id_raw)
            except (TypeError, ValueError):
                receta_id = None
                fila_errores.append('Receta ID inválido')
            if receta_id:
                receta_plan = recetas_preview['plans_by_id'].get(receta_id)
                receta_obj = recetas_por_id.get(receta_id)
                if not receta_plan and not receta_obj:
                    fila_errores.append('No se encontró una receta accesible con ese ID')
        elif receta_lookup:
            receta_plan = recetas_preview['plans_by_name'].get(receta_lookup)
            receta_obj = recetas_por_nombre.get(receta_lookup)
            if receta_obj is None and receta_lookup in recetas_por_nombre:
                fila_errores.append('Hay más de una receta existente con ese nombre; use Receta ID')
            elif not receta_plan and not receta_obj:
                fila_errores.append('No se encontró una receta con ese nombre')
        else:
            fila_errores.append('Debe informar Receta ID o Receta Nombre')

        producto = None
        producto_label = str(producto_nombre_raw).strip() if producto_nombre_raw is not None else ''
        if producto_id_raw is not None and str(producto_id_raw).strip() != '':
            try:
                producto_id = _parse_excel_int(producto_id_raw)
            except (TypeError, ValueError):
                producto_id = None
                fila_errores.append('Producto ID inválido')
            if producto_id:
                producto = productos_por_id.get(producto_id)
                if not producto:
                    fila_errores.append('No se encontró un producto accesible con ese ID')
        else:
            producto_lookup = _normalize_lookup_text(producto_nombre_raw)
            if not producto_lookup:
                fila_errores.append('Debe informar Producto ID o Producto Nombre')
            else:
                producto = productos_por_nombre.get(producto_lookup)
                if producto is None and producto_lookup in productos_por_nombre:
                    fila_errores.append('Hay más de un producto con ese nombre; use Producto ID')
                elif not producto:
                    fila_errores.append('No se encontró un producto con ese nombre')

        try:
            cantidad = _parse_excel_decimal(
                cantidad_raw,
                min_value=Decimal('0.01'),
            )
        except (InvalidOperation, ValueError):
            fila_errores.append('Cantidad debe ser un número mayor a 0')
            cantidad = Decimal('1')

        medida_uso = str(medida_raw).strip() if medida_raw is not None else ''
        if not medida_uso:
            fila_errores.append('Medida de uso es obligatoria')
        elif medida_uso not in UNIDADES_VALIDAS:
            fila_errores.append('Medida de uso inválida')
        elif producto:
            try:
                _validate_medida_producto_receta(producto, medida_uso)
            except ValueError as exc:
                fila_errores.append(str(exc))

        if relacion_id and relacion_id not in relaciones_por_id:
            fila_errores.append('No existe una relación accesible con ese ID')

        recipe_anchor = None
        recipe_label = ''
        recipe_existing_id = None
        if receta_plan:
            recipe_anchor = receta_plan['plan_key']
            recipe_label = receta_plan['nombre']
            recipe_existing_id = receta_plan['existing_id']
        elif receta_obj:
            recipe_anchor = f"db-{receta_obj.pk}"
            recipe_label = receta_obj.nombre
            recipe_existing_id = receta_obj.pk

        dedupe_key = None
        existing_relation = None
        if relacion_id:
            existing_relation = relaciones_por_id[relacion_id]
            dedupe_key = ('id', relacion_id)
        elif recipe_anchor and producto:
            dedupe_key = ('combo', recipe_anchor, producto.pk)
            if recipe_existing_id:
                existing_relation = relaciones_qs.filter(receta_id=recipe_existing_id, producto_id=producto.pk).first()

        if dedupe_key in seen_relation_keys:
            fila_errores.append(
                f"La relación ya fue incluida en la fila {seen_relation_keys[dedupe_key]} de la hoja {HOJA_RECETAS_PRODUCTOS}"
            )

        if fila_errores:
            errores.append(f"Hoja {HOJA_RECETAS_PRODUCTOS}, fila {row_num}: " + ' | '.join(fila_errores))
            continue

        seen_relation_keys[dedupe_key] = row_num
        action = 'actualizar' if existing_relation else 'crear'
        plan = {
            'row_num': row_num,
            'action': action,
            'existing_id': existing_relation.pk if existing_relation else None,
            'recipe_plan_key': receta_plan['plan_key'] if receta_plan else None,
            'recipe_existing_id': recipe_existing_id,
            'product_id': producto.pk,
            'cantidad': str(cantidad),
            'medida_uso': medida_uso,
        }
        planes.append(plan)
        acciones.append({
            'sheet': HOJA_RECETAS_PRODUCTOS,
            'row_num': row_num,
            'label': f"{recipe_label} -> {producto_label or producto.nombre}",
            'action': _preview_action_label(action),
        })

    return {
        'errors': errores,
        'plans': planes,
        'actions': acciones,
        'total_rows': total_filas,
    }


def _leer_hoja_recetas_bienes(ws, recetas_qs, bienes_qs, recetas_preview, bienes_preview, request):
    errores = []
    planes = []
    acciones = []
    total_filas = 0

    bienes_por_id = {bien.pk: bien for bien in bienes_qs}
    bienes_por_nombre = _build_name_lookup(bienes_qs)
    recetas_por_id = {receta.pk: receta for receta in recetas_qs}
    recetas_por_nombre = _build_name_lookup(recetas_qs)
    relaciones_qs = BienReceta.objects.select_related('receta', 'bien')
    if not request.user.is_superuser:
        relaciones_qs = relaciones_qs.filter(receta__usuario=request.user.username)

    relaciones_por_id = {relacion.pk: relacion for relacion in relaciones_qs}
    seen_relation_keys = {}

    encabezados = _worksheet_headers(ws, COLUMNAS_RECETA_BIEN)
    esperados = [header for _, header in COLUMNAS_RECETA_BIEN]
    if encabezados != esperados:
        errores.append(
            f"Hoja {HOJA_RECETAS_BIENES}: encabezados inválidos. Se esperaba: {', '.join(esperados)}."
        )
        return {
            'errors': errores,
            'plans': planes,
            'actions': acciones,
            'total_rows': total_filas,
        }

    for row_num, row in _iter_sheet_rows(ws, COLUMNAS_RECETA_BIEN):
        total_filas += 1
        id_raw, receta_id_raw, receta_nombre_raw, bien_id_raw, bien_nombre_raw, tiempo_cantidad_raw, tiempo_unidad_raw, incluir_depreciacion_raw, incluir_electricidad_raw, observaciones_raw = row
        fila_errores = []

        try:
            relacion_id = _parse_excel_int(id_raw) if id_raw is not None and str(id_raw).strip() != '' else None
        except (TypeError, ValueError):
            relacion_id = None
            fila_errores.append('ID de relación inválido')

        receta_plan = None
        receta_obj = None
        receta_lookup = _normalize_lookup_text(receta_nombre_raw)
        if receta_id_raw is not None and str(receta_id_raw).strip() != '':
            try:
                receta_id = _parse_excel_int(receta_id_raw)
            except (TypeError, ValueError):
                receta_id = None
                fila_errores.append('Receta ID inválido')
            if receta_id:
                receta_plan = recetas_preview['plans_by_id'].get(receta_id)
                receta_obj = recetas_por_id.get(receta_id)
                if not receta_plan and not receta_obj:
                    fila_errores.append('No se encontró una receta accesible con ese ID')
        elif receta_lookup:
            receta_plan = recetas_preview['plans_by_name'].get(receta_lookup)
            receta_obj = recetas_por_nombre.get(receta_lookup)
            if receta_obj is None and receta_lookup in recetas_por_nombre:
                fila_errores.append('Hay más de una receta existente con ese nombre; use Receta ID')
            elif not receta_plan and not receta_obj:
                fila_errores.append('No se encontró una receta con ese nombre')
        else:
            fila_errores.append('Debe informar Receta ID o Receta Nombre')

        bien_plan = None
        bien_obj = None
        bien_lookup = _normalize_lookup_text(bien_nombre_raw)
        bien_label = str(bien_nombre_raw).strip() if bien_nombre_raw is not None else ''
        if bien_id_raw is not None and str(bien_id_raw).strip() != '':
            try:
                bien_id = _parse_excel_int(bien_id_raw)
            except (TypeError, ValueError):
                bien_id = None
                fila_errores.append('Bien ID inválido')
            if bien_id:
                bien_plan = bienes_preview['plans_by_id'].get(bien_id)
                bien_obj = bienes_por_id.get(bien_id)
                if not bien_plan and not bien_obj:
                    fila_errores.append('No se encontró un bien accesible con ese ID')
        elif bien_lookup:
            bien_plan = bienes_preview['plans_by_name'].get(bien_lookup)
            bien_obj = bienes_por_nombre.get(bien_lookup)
            if bien_obj is None and bien_lookup in bienes_por_nombre:
                fila_errores.append('Hay más de un bien existente con ese nombre; use Bien ID')
            elif not bien_plan and not bien_obj:
                fila_errores.append('No se encontró un bien con ese nombre')
        else:
            fila_errores.append('Debe informar Bien ID o Bien Nombre')

        try:
            tiempo_cantidad = _parse_excel_decimal(tiempo_cantidad_raw, min_value=Decimal('0.01'))
        except (InvalidOperation, ValueError):
            fila_errores.append('Tiempo de uso cantidad debe ser un número mayor a 0')
            tiempo_cantidad = Decimal('1')

        tiempo_unidad = str(tiempo_unidad_raw).strip() if tiempo_unidad_raw is not None else ''
        if not tiempo_unidad:
            fila_errores.append('Tiempo de uso unidad es obligatoria')
        elif tiempo_unidad not in UNIDADES_TIEMPO_VALIDAS:
            fila_errores.append('Tiempo de uso unidad inválida')

        try:
            incluir_depreciacion = _parse_excel_bool(incluir_depreciacion_raw, default=True)
        except ValueError:
            fila_errores.append("Incluir depreciación debe ser 'Si' o 'No'")
            incluir_depreciacion = True

        try:
            incluir_electricidad = _parse_excel_bool(incluir_electricidad_raw, default=True)
        except ValueError:
            fila_errores.append("Incluir electricidad debe ser 'Si' o 'No'")
            incluir_electricidad = True

        observaciones = str(observaciones_raw).strip() if observaciones_raw is not None else ''

        if relacion_id and relacion_id not in relaciones_por_id:
            fila_errores.append('No existe una relación accesible con ese ID')

        recipe_anchor = None
        recipe_label = ''
        recipe_existing_id = None
        if receta_plan:
            recipe_anchor = receta_plan['plan_key']
            recipe_label = receta_plan['nombre']
            recipe_existing_id = receta_plan['existing_id']
        elif receta_obj:
            recipe_anchor = f"db-receta-{receta_obj.pk}"
            recipe_label = receta_obj.nombre
            recipe_existing_id = receta_obj.pk

        bien_anchor = None
        bien_existing_id = None
        if bien_plan:
            bien_anchor = bien_plan['plan_key']
            bien_existing_id = bien_plan['existing_id']
        elif bien_obj:
            bien_anchor = f"db-bien-{bien_obj.pk}"
            bien_existing_id = bien_obj.pk

        dedupe_key = None
        existing_relation = None
        if relacion_id:
            existing_relation = relaciones_por_id[relacion_id]
            dedupe_key = ('id', relacion_id)
        elif recipe_anchor and bien_anchor:
            dedupe_key = ('combo', recipe_anchor, bien_anchor)
            if recipe_existing_id and bien_existing_id:
                existing_relation = relaciones_qs.filter(receta_id=recipe_existing_id, bien_id=bien_existing_id).first()

        if dedupe_key in seen_relation_keys:
            fila_errores.append(
                f"La relación ya fue incluida en la fila {seen_relation_keys[dedupe_key]} de la hoja {HOJA_RECETAS_BIENES}"
            )

        if fila_errores:
            errores.append(f"Hoja {HOJA_RECETAS_BIENES}, fila {row_num}: " + ' | '.join(fila_errores))
            continue

        seen_relation_keys[dedupe_key] = row_num
        action = 'actualizar' if existing_relation else 'crear'
        plan = {
            'row_num': row_num,
            'action': action,
            'existing_id': existing_relation.pk if existing_relation else None,
            'recipe_plan_key': receta_plan['plan_key'] if receta_plan else None,
            'recipe_existing_id': recipe_existing_id,
            'bien_plan_key': bien_plan['plan_key'] if bien_plan else None,
            'bien_existing_id': bien_existing_id,
            'tiempo_uso_cantidad': str(tiempo_cantidad),
            'tiempo_uso_unidad': tiempo_unidad,
            'incluir_depreciacion': incluir_depreciacion,
            'incluir_electricidad': incluir_electricidad,
            'observaciones': observaciones,
        }
        planes.append(plan)
        acciones.append({
            'sheet': HOJA_RECETAS_BIENES,
            'row_num': row_num,
            'label': f"{recipe_label} -> {bien_label or (bien_obj.nombre if bien_obj else bien_plan['nombre'])}",
            'action': _preview_action_label(action),
        })

    return {
        'errors': errores,
        'plans': planes,
        'actions': acciones,
        'total_rows': total_filas,
    }


def _analizar_archivo_recetas(workbook_path, request):
    preview = {
        'token': uuid.uuid4().hex,
        'errors': [],
        'warnings': [],
        'summary': {
            'configuraciones_crear': 0,
            'configuraciones_actualizar': 0,
            'recetas_crear': 0,
            'recetas_actualizar': 0,
            'bienes_crear': 0,
            'bienes_actualizar': 0,
            'relaciones_crear': 0,
            'relaciones_actualizar': 0,
            'relaciones_bienes_crear': 0,
            'relaciones_bienes_actualizar': 0,
            'configuracion_filas': 0,
            'recetas_filas': 0,
            'bienes_filas': 0,
            'ingredientes_filas': 0,
            'bienes_receta_filas': 0,
        },
        'configuracion_actions': [],
        'receta_actions': [],
        'bien_actions': [],
        'relacion_actions': [],
        'bien_relacion_actions': [],
        'high_volume': False,
        'config_plan': None,
        'recipe_plans': [],
        'bien_plans': [],
        'relation_plans': [],
        'bien_relation_plans': [],
    }

    wb = None
    try:
        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        preview['errors'].append(f'No se pudo leer el archivo: {exc}')
        return preview
    try:
        required_sheets = [
            HOJA_RECETAS,
            HOJA_RECETAS_PRODUCTOS,
            HOJA_BIENES,
            HOJA_RECETAS_BIENES,
            HOJA_CONFIGURACION_RECETAS,
            HOJA_PRODUCTOS,
        ]
        missing_sheets = [sheet for sheet in required_sheets if sheet not in wb.sheetnames]
        if missing_sheets:
            preview['errors'].append(
                'Faltan hojas obligatorias: ' + ', '.join(missing_sheets)
            )
            return preview

        usuario = request.user.username
        if request.user.is_superuser:
            recetas_qs = list(Receta.objects.select_related('categoria').all())
            productos_qs = list(Producto.objects.select_related('categoria').all())
            bienes_qs = list(Bien.objects.all())
        else:
            recetas_qs = list(Receta.objects.select_related('categoria').filter(usuario=usuario))
            productos_qs = list(Producto.objects.select_related('categoria').filter(usuario=usuario))
            bienes_qs = list(Bien.objects.filter(usuario=usuario))

        productos_headers = _worksheet_headers(wb[HOJA_PRODUCTOS], COLUMNAS_PRODUCTOS_REFERENCIA)
        productos_esperados = [header for _, header in COLUMNAS_PRODUCTOS_REFERENCIA]
        if productos_headers != productos_esperados:
            preview['errors'].append(
                f"Hoja {HOJA_PRODUCTOS}: encabezados inválidos. Se esperaba: {', '.join(productos_esperados)}."
            )
            return preview

        configuracion_preview = _leer_hoja_configuracion_recetas(wb[HOJA_CONFIGURACION_RECETAS])
        recetas_preview = _leer_hoja_recetas(wb[HOJA_RECETAS], recetas_qs, request)
        bienes_preview = _leer_hoja_bienes(wb[HOJA_BIENES], bienes_qs, request)
        relaciones_preview = _leer_hoja_recetas_productos(
            wb[HOJA_RECETAS_PRODUCTOS],
            recetas_qs,
            productos_qs,
            recetas_preview,
            request,
        )
        relaciones_bienes_preview = _leer_hoja_recetas_bienes(
            wb[HOJA_RECETAS_BIENES],
            recetas_qs,
            bienes_qs,
            recetas_preview,
            bienes_preview,
            request,
        )

        preview['errors'].extend(configuracion_preview['errors'])
        preview['warnings'].extend(configuracion_preview['warnings'])
        preview['errors'].extend(recetas_preview['errors'])
        preview['errors'].extend(bienes_preview['errors'])
        preview['errors'].extend(relaciones_preview['errors'])
        preview['errors'].extend(relaciones_bienes_preview['errors'])
        preview['warnings'].extend(recetas_preview['warnings'])
        preview['warnings'].extend(bienes_preview['warnings'])
        preview['summary']['configuracion_filas'] = configuracion_preview['total_rows']
        preview['summary']['recetas_filas'] = recetas_preview['total_rows']
        preview['summary']['bienes_filas'] = bienes_preview['total_rows']
        preview['summary']['ingredientes_filas'] = relaciones_preview['total_rows']
        preview['summary']['bienes_receta_filas'] = relaciones_bienes_preview['total_rows']
        preview['summary']['configuraciones_crear'] = 1 if configuracion_preview['plan'] and configuracion_preview['plan']['action'] == 'crear' else 0
        preview['summary']['configuraciones_actualizar'] = 1 if configuracion_preview['plan'] and configuracion_preview['plan']['action'] == 'actualizar' else 0
        preview['summary']['recetas_crear'] = sum(1 for plan in recetas_preview['plans'] if plan['action'] == 'crear')
        preview['summary']['recetas_actualizar'] = sum(1 for plan in recetas_preview['plans'] if plan['action'] == 'actualizar')
        preview['summary']['bienes_crear'] = sum(1 for plan in bienes_preview['plans'] if plan['action'] == 'crear')
        preview['summary']['bienes_actualizar'] = sum(1 for plan in bienes_preview['plans'] if plan['action'] == 'actualizar')
        preview['summary']['relaciones_crear'] = sum(1 for plan in relaciones_preview['plans'] if plan['action'] == 'crear')
        preview['summary']['relaciones_actualizar'] = sum(1 for plan in relaciones_preview['plans'] if plan['action'] == 'actualizar')
        preview['summary']['relaciones_bienes_crear'] = sum(1 for plan in relaciones_bienes_preview['plans'] if plan['action'] == 'crear')
        preview['summary']['relaciones_bienes_actualizar'] = sum(1 for plan in relaciones_bienes_preview['plans'] if plan['action'] == 'actualizar')
        preview['configuracion_actions'] = configuracion_preview['actions']
        preview['receta_actions'] = recetas_preview['actions']
        preview['bien_actions'] = bienes_preview['actions']
        preview['relacion_actions'] = relaciones_preview['actions']
        preview['bien_relacion_actions'] = relaciones_bienes_preview['actions']
        preview['config_plan'] = configuracion_preview['plan']
        preview['recipe_plans'] = recetas_preview['plans']
        preview['bien_plans'] = bienes_preview['plans']
        preview['relation_plans'] = relaciones_preview['plans']
        preview['bien_relation_plans'] = relaciones_bienes_preview['plans']

        if (
            preview['summary']['recetas_filas'] >= PREVIEW_RECETAS_WARN_UMBRAL
            or preview['summary']['ingredientes_filas'] >= PREVIEW_INGREDIENTES_WARN_UMBRAL
            or preview['summary']['bienes_filas'] >= PREVIEW_BIENES_WARN_UMBRAL
            or preview['summary']['bienes_receta_filas'] >= PREVIEW_BIENES_WARN_UMBRAL
        ):
            preview['high_volume'] = True
            preview['warnings'].append(
                'El archivo tiene un volumen alto. En una versión corporativa conviene mover la confirmación a un proceso asíncrono para evitar timeouts web.'
            )

        if (
            not preview['config_plan']
            and not preview['recipe_plans']
            and not preview['bien_plans']
            and not preview['relation_plans']
            and not preview['bien_relation_plans']
            and not preview['errors']
        ):
            preview['warnings'].append('No se encontraron filas con datos para analizar.')

        return preview
    finally:
        if wb is not None:
            wb.close()


def _build_workbook_recetas(productos, recetas=None, relaciones=None, bienes=None, relaciones_bienes=None, configuracion=None, solo_plantilla=False):
    wb = openpyxl.Workbook()

    bienes = list(bienes or [])
    relaciones_bienes = list(relaciones_bienes or [])

    ws_listas = wb.active
    ws_listas.title = HOJA_LISTAS_RECETAS
    ws_listas.sheet_state = 'hidden'

    for row_idx, producto in enumerate(productos, start=1):
        ws_listas.cell(row=row_idx, column=1, value=producto.pk)
        ws_listas.cell(row=row_idx, column=2, value=producto.nombre)

    for row_idx, unidad in enumerate(UNIDADES_VALIDAS, start=1):
        ws_listas.cell(row=row_idx, column=4, value=unidad)

    for row_idx, bien in enumerate(bienes, start=1):
        ws_listas.cell(row=row_idx, column=6, value=bien.pk)
        ws_listas.cell(row=row_idx, column=7, value=bien.nombre)

    for row_idx, unidad in enumerate(UNIDADES_TIEMPO_VALIDAS, start=1):
        ws_listas.cell(row=row_idx, column=9, value=unidad)

    for row_idx, opcion in enumerate(OPCIONES_BOOLEANAS_IMPORTACION, start=1):
        ws_listas.cell(row=row_idx, column=11, value=opcion)

    ws_recetas = wb.create_sheet(title=HOJA_RECETAS)
    ws_relaciones = wb.create_sheet(title=HOJA_RECETAS_PRODUCTOS)
    ws_bienes = wb.create_sheet(title=HOJA_BIENES)
    ws_relaciones_bienes = wb.create_sheet(title=HOJA_RECETAS_BIENES)
    ws_configuracion = wb.create_sheet(title=HOJA_CONFIGURACION_RECETAS)
    ws_productos = wb.create_sheet(title=HOJA_PRODUCTOS)

    for col_idx, (_, header) in enumerate(COLUMNAS_RECETA_IMPORTACION, start=1):
        _apply_header_style_receta(ws_recetas.cell(row=1, column=col_idx, value=header))
    for col_idx, (_, header) in enumerate(COLUMNAS_RECETA_PRODUCTO, start=1):
        _apply_header_style_receta(ws_relaciones.cell(row=1, column=col_idx, value=header))
    for col_idx, (_, header) in enumerate(COLUMNAS_BIEN_IMPORTACION, start=1):
        _apply_header_style_receta(ws_bienes.cell(row=1, column=col_idx, value=header))
    for col_idx, (_, header) in enumerate(COLUMNAS_RECETA_BIEN, start=1):
        _apply_header_style_receta(ws_relaciones_bienes.cell(row=1, column=col_idx, value=header))
    for col_idx, (_, header) in enumerate(COLUMNAS_CONFIGURACION_RECETAS, start=1):
        _apply_header_style_receta(ws_configuracion.cell(row=1, column=col_idx, value=header))
    for col_idx, (_, header) in enumerate(COLUMNAS_PRODUCTOS_REFERENCIA, start=1):
        _apply_header_style_receta(ws_productos.cell(row=1, column=col_idx, value=header), readonly=True)

    for col_idx, ancho in enumerate([10, 34, 34, 24, 14, 18, 12, 34], start=1):
        ws_recetas.column_dimensions[get_column_letter(col_idx)].width = ancho
    for col_idx, ancho in enumerate([10, 12, 32, 12, 32, 14, 18], start=1):
        ws_relaciones.column_dimensions[get_column_letter(col_idx)].width = ancho
    for col_idx, ancho in enumerate([10, 30, 34, 16, 18, 18, 16, 16, 12], start=1):
        ws_bienes.column_dimensions[get_column_letter(col_idx)].width = ancho
    for col_idx, ancho in enumerate([10, 12, 32, 12, 32, 18, 18, 20, 20, 30], start=1):
        ws_relaciones_bienes.column_dimensions[get_column_letter(col_idx)].width = ancho
    ws_configuracion.column_dimensions['A'].width = 18
    for col_idx, ancho in enumerate([10, 30, 20, 24, 20, 12, 14], start=1):
        ws_productos.column_dimensions[get_column_letter(col_idx)].width = ancho

    for worksheet in [ws_recetas, ws_relaciones, ws_bienes, ws_relaciones_bienes, ws_configuracion, ws_productos]:
        worksheet.row_dimensions[1].height = 28
        worksheet.freeze_panes = 'A2'

    fill_example = PatternFill(fill_type='solid', fgColor='F0F0F0')
    font_example = Font(italic=True, color='888888')
    fill_id = PatternFill(fill_type='solid', fgColor='E8E8E8')
    font_id = Font(bold=True, color='555555')
    fill_par = PatternFill(fill_type='solid', fgColor='F7F9FC')
    fill_readonly = PatternFill(fill_type='solid', fgColor='EFEFEF')
    font_readonly = Font(italic=True, color='888888')

    product_formula_id = f"'{HOJA_LISTAS_RECETAS}'!$A$1:$A${max(len(productos), 1)}"
    product_formula_name = f"'{HOJA_LISTAS_RECETAS}'!$B$1:$B${max(len(productos), 1)}"
    unidad_formula = f"'{HOJA_LISTAS_RECETAS}'!$D$1:$D${len(UNIDADES_VALIDAS)}"
    bien_formula_id = f"'{HOJA_LISTAS_RECETAS}'!$F$1:$F${max(len(bienes), 1)}"
    bien_formula_name = f"'{HOJA_LISTAS_RECETAS}'!$G$1:$G${max(len(bienes), 1)}"
    tiempo_formula = f"'{HOJA_LISTAS_RECETAS}'!$I$1:$I${len(UNIDADES_TIEMPO_VALIDAS)}"
    boolean_formula = f"'{HOJA_LISTAS_RECETAS}'!$K$1:$K${len(OPCIONES_BOOLEANAS_IMPORTACION)}"

    dv_producto_id = DataValidation(type='list', formula1=product_formula_id, allow_blank=True)
    dv_producto_nombre = DataValidation(type='list', formula1=product_formula_name, allow_blank=True)
    dv_unidad = DataValidation(type='list', formula1=unidad_formula, allow_blank=False)
    ws_relaciones.add_data_validation(dv_producto_id)
    ws_relaciones.add_data_validation(dv_producto_nombre)
    ws_relaciones.add_data_validation(dv_unidad)
    dv_producto_id.sqref = 'D2:D10000'
    dv_producto_nombre.sqref = 'E2:E10000'
    dv_unidad.sqref = 'G2:G10000'

    dv_bien_id = DataValidation(type='list', formula1=bien_formula_id, allow_blank=True)
    dv_bien_nombre = DataValidation(type='list', formula1=bien_formula_name, allow_blank=True)
    dv_tiempo_relacion = DataValidation(type='list', formula1=tiempo_formula, allow_blank=False)
    dv_bool_relacion = DataValidation(type='list', formula1=boolean_formula, allow_blank=True)
    dv_tiempo_bien = DataValidation(type='list', formula1=tiempo_formula, allow_blank=False)
    dv_activo = DataValidation(type='list', formula1=boolean_formula, allow_blank=True)
    ws_relaciones_bienes.add_data_validation(dv_bien_id)
    ws_relaciones_bienes.add_data_validation(dv_bien_nombre)
    ws_relaciones_bienes.add_data_validation(dv_tiempo_relacion)
    ws_relaciones_bienes.add_data_validation(dv_bool_relacion)
    ws_bienes.add_data_validation(dv_tiempo_bien)
    ws_bienes.add_data_validation(dv_activo)
    dv_bien_id.sqref = 'D2:D10000'
    dv_bien_nombre.sqref = 'E2:E10000'
    dv_tiempo_relacion.sqref = 'G2:G10000'
    dv_bool_relacion.sqref = 'H2:I10000'
    dv_tiempo_bien.sqref = 'F2:F10000'
    dv_activo.sqref = 'I2:I10000'

    if solo_plantilla:
        producto_ejemplo = productos[0] if productos else None
        bien_ejemplo = bienes[0] if bienes else None
        ejemplo_receta = ['', 'Torta de Chocolate', 'Torta esponjosa con relleno', 'Tortas', '12', '50', '21', 'Receta especial']
        ejemplo_relacion = ['', '', 'Torta de Chocolate', producto_ejemplo.pk if producto_ejemplo else '', producto_ejemplo.nombre if producto_ejemplo else '', '0.50', producto_ejemplo.unidad_de_medida if producto_ejemplo else 'Gramos']
        ejemplo_bien = ['', 'Horno convector', 'Equipo principal para cocción', '850000', '6000', 'Horas', '2200', '60', 'Si']
        ejemplo_relacion_bien = ['', '', 'Torta de Chocolate', bien_ejemplo.pk if bien_ejemplo else '', bien_ejemplo.nombre if bien_ejemplo else 'Horno convector', '40', 'Minutos', 'Si', 'Si', 'Uso principal de cocción']
        ejemplos = [
            (ws_recetas, ejemplo_receta),
            (ws_relaciones, ejemplo_relacion),
            (ws_bienes, ejemplo_bien),
            (ws_relaciones_bienes, ejemplo_relacion_bien),
        ]
        for worksheet, valores in ejemplos:
            for col_idx, valor in enumerate(valores, start=1):
                cell = worksheet.cell(row=2, column=col_idx, value=valor)
                cell.fill = fill_example
                cell.font = font_example
        config_cell = ws_configuracion.cell(row=2, column=1, value=125.50)
        config_cell.fill = fill_example
        config_cell.font = font_example
    else:
        for row_idx, receta in enumerate(recetas or [], start=2):
            fill = fill_par if row_idx % 2 == 0 else None
            valores = [receta.pk, receta.nombre or '', receta.descripcion or '', receta.categoria.nombre if receta.categoria else '', _safe_float(receta.porciones) if receta.porciones is not None else 1, _safe_float(receta.rentabilidad) if receta.rentabilidad is not None else 0, _safe_float(receta.iva) if receta.iva is not None else 21, receta.comentarios or '']
            for col_idx, valor in enumerate(valores, start=1):
                cell = ws_recetas.cell(row=row_idx, column=col_idx, value=valor)
                if col_idx == 1:
                    cell.fill = fill_id
                    cell.font = font_id
                elif fill:
                    cell.fill = fill

        for row_idx, relacion in enumerate(relaciones or [], start=2):
            fill = fill_par if row_idx % 2 == 0 else None
            valores = [relacion.pk, relacion.receta_id, relacion.receta.nombre if relacion.receta else '', relacion.producto_id, relacion.producto.nombre if relacion.producto else '', _safe_float(relacion.cantidad) if relacion.cantidad is not None else 1, relacion.medida_uso or '']
            for col_idx, valor in enumerate(valores, start=1):
                cell = ws_relaciones.cell(row=row_idx, column=col_idx, value=valor)
                if col_idx == 1:
                    cell.fill = fill_id
                    cell.font = font_id
                elif fill:
                    cell.fill = fill

        for row_idx, bien in enumerate(bienes, start=2):
            fill = fill_par if row_idx % 2 == 0 else None
            valores = [bien.pk, bien.nombre or '', bien.descripcion or '', _safe_float(bien.costo_compra) if bien.costo_compra is not None else 0, _safe_float(bien.vida_util_cantidad) if bien.vida_util_cantidad is not None else 1, bien.vida_util_unidad or 'Horas', _safe_float(bien.potencia_watts) if bien.potencia_watts is not None else 0, _safe_float(bien.factor_uso_porcentaje) if bien.factor_uso_porcentaje is not None else 100, 'Si' if bien.activo else 'No']
            for col_idx, valor in enumerate(valores, start=1):
                cell = ws_bienes.cell(row=row_idx, column=col_idx, value=valor)
                if col_idx == 1:
                    cell.fill = fill_id
                    cell.font = font_id
                elif fill:
                    cell.fill = fill

        for row_idx, relacion in enumerate(relaciones_bienes, start=2):
            fill = fill_par if row_idx % 2 == 0 else None
            valores = [relacion.pk, relacion.receta_id, relacion.receta.nombre if relacion.receta else '', relacion.bien_id, relacion.bien.nombre if relacion.bien else '', _safe_float(relacion.tiempo_uso_cantidad) if relacion.tiempo_uso_cantidad is not None else 1, relacion.tiempo_uso_unidad or 'Minutos', 'Si' if relacion.incluir_depreciacion else 'No', 'Si' if relacion.incluir_electricidad else 'No', relacion.observaciones or '']
            for col_idx, valor in enumerate(valores, start=1):
                cell = ws_relaciones_bienes.cell(row=row_idx, column=col_idx, value=valor)
                if col_idx == 1:
                    cell.fill = fill_id
                    cell.font = font_id
                elif fill:
                    cell.fill = fill

        if configuracion:
            ws_configuracion.cell(row=2, column=1, value=_safe_float(configuracion.precio_kwh))

    for row_idx, producto in enumerate(productos, start=2):
        fill = fill_par if row_idx % 2 == 0 else None
        valores = [producto.pk, producto.nombre or '', producto.unidad_de_medida or '', producto.categoria.nombre if producto.categoria else '', producto.marca or '', _safe_float(producto.cantidad) if producto.cantidad is not None else 1, _safe_float(producto.costo) if producto.costo is not None else 0]
        for col_idx, valor in enumerate(valores, start=1):
            cell = ws_productos.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill = fill_readonly if col_idx != 1 else fill_id
            cell.font = font_readonly if col_idx != 1 else font_id
            if fill and col_idx == 1:
                cell.fill = fill_id

    return wb


def _confirmar_importacion_recetas(preview, request):
    analysis = _analizar_archivo_recetas(preview['temp_path'], request)
    analysis['token'] = preview['token']
    if analysis['errors']:
        return False, analysis

    usuario = request.user.username
    recetas_guardadas = {}
    bienes_guardados = {}

    with transaction.atomic():
        if analysis['config_plan']:
            config_plan = analysis['config_plan']
            configuracion = None
            if config_plan['existing_id']:
                configuracion = Configuracion.objects.filter(pk=config_plan['existing_id']).first()
            if not configuracion:
                configuracion = Configuracion.objects.first()
            if not configuracion:
                configuracion = Configuracion(
                    nombre_emprendimiento=f'Configuración {usuario}',
                    moneda='$',
                    redondeo=Decimal('0'),
                    usuario=usuario,
                )
            if not configuracion.nombre_emprendimiento:
                configuracion.nombre_emprendimiento = f'Configuración {usuario}'
            if not configuracion.moneda:
                configuracion.moneda = '$'
            if configuracion.redondeo is None:
                configuracion.redondeo = Decimal('0')
            if not configuracion.usuario:
                configuracion.usuario = usuario
            configuracion.precio_kwh = Decimal(config_plan['precio_kwh'])
            configuracion.save()

        for plan in analysis['recipe_plans']:
            receta = None
            if plan['existing_id']:
                receta = Receta.objects.get(pk=plan['existing_id'])
            if not receta:
                receta = Receta(usuario=plan['owner_user'])

            receta.nombre = plan['nombre']
            receta.descripcion = plan['descripcion']
            receta.categoria = _resolve_categoria_receta(plan['categoria'], plan['owner_user']) if plan['categoria'] else None
            receta.porciones = Decimal(plan['porciones'])
            receta.rentabilidad = Decimal(plan['rentabilidad'])
            receta.iva = Decimal(plan['iva'])
            receta.comentarios = plan['comentarios'] or None
            if not receta.usuario:
                receta.usuario = plan['owner_user'] or usuario
            receta.save()
            recetas_guardadas[plan['plan_key']] = receta
            if receta.pk and plan['existing_id']:
                recetas_guardadas[f"db-{receta.pk}"] = receta

        for plan in analysis['bien_plans']:
            bien = None
            if plan['existing_id']:
                bien = Bien.objects.get(pk=plan['existing_id'])
            if not bien:
                bien = Bien(usuario=plan['owner_user'])

            bien.nombre = plan['nombre']
            bien.descripcion = plan['descripcion'] or None
            bien.costo_compra = Decimal(plan['costo_compra'])
            bien.vida_util_cantidad = Decimal(plan['vida_util_cantidad'])
            bien.vida_util_unidad = plan['vida_util_unidad']
            bien.potencia_watts = Decimal(plan['potencia_watts'])
            bien.factor_uso_porcentaje = Decimal(plan['factor_uso_porcentaje'])
            bien.activo = bool(plan['activo'])
            if not bien.usuario:
                bien.usuario = plan['owner_user'] or usuario
            bien.save()
            bienes_guardados[plan['plan_key']] = bien
            if bien.pk and plan['existing_id']:
                bienes_guardados[f"db-{bien.pk}"] = bien

        for plan in analysis['relation_plans']:
            receta = None
            if plan['recipe_plan_key']:
                receta = recetas_guardadas.get(plan['recipe_plan_key'])
            if not receta and plan['recipe_existing_id']:
                receta = recetas_guardadas.get(f"db-{plan['recipe_existing_id']}")
            if not receta and plan['recipe_existing_id']:
                receta = Receta.objects.get(pk=plan['recipe_existing_id'])

            producto = Producto.objects.get(pk=plan['product_id'])
            relacion = None
            if plan['existing_id']:
                relacion = ProductoReceta.objects.get(pk=plan['existing_id'])
            if not relacion:
                relacion = ProductoReceta(receta=receta, producto=producto)

            relacion.receta = receta
            relacion.producto = producto
            relacion.cantidad = Decimal(plan['cantidad'])
            relacion.medida_uso = plan['medida_uso']
            relacion.usuario = receta.usuario
            relacion.save()

        for plan in analysis['bien_relation_plans']:
            receta = None
            if plan['recipe_plan_key']:
                receta = recetas_guardadas.get(plan['recipe_plan_key'])
            if not receta and plan['recipe_existing_id']:
                receta = recetas_guardadas.get(f"db-{plan['recipe_existing_id']}")
            if not receta and plan['recipe_existing_id']:
                receta = Receta.objects.get(pk=plan['recipe_existing_id'])

            bien = None
            if plan['bien_plan_key']:
                bien = bienes_guardados.get(plan['bien_plan_key'])
            if not bien and plan['bien_existing_id']:
                bien = bienes_guardados.get(f"db-{plan['bien_existing_id']}")
            if not bien and plan['bien_existing_id']:
                bien = Bien.objects.get(pk=plan['bien_existing_id'])

            relacion = None
            if plan['existing_id']:
                relacion = BienReceta.objects.get(pk=plan['existing_id'])
            if not relacion:
                relacion = BienReceta(receta=receta, bien=bien)

            relacion.receta = receta
            relacion.bien = bien
            relacion.tiempo_uso_cantidad = Decimal(plan['tiempo_uso_cantidad'])
            relacion.tiempo_uso_unidad = plan['tiempo_uso_unidad']
            relacion.incluir_depreciacion = bool(plan['incluir_depreciacion'])
            relacion.incluir_electricidad = bool(plan['incluir_electricidad'])
            relacion.observaciones = plan['observaciones'] or None
            relacion.usuario = receta.usuario
            relacion.save()

    return True, analysis


@staff_member_required
def importar_exportar_recetas(request):
    """Vista principal para importación / exportación masiva de recetas."""

    usuario = request.user.username
    config = Configuracion.objects.filter(usuario=usuario).first()

    if request.user.is_superuser:
        recetas_qs = Receta.objects.select_related('categoria').all()
        productos_qs = Producto.objects.select_related('categoria').all()
        relaciones_qs = ProductoReceta.objects.select_related('receta', 'producto').all()
        bienes_qs = Bien.objects.all()
        relaciones_bienes_qs = BienReceta.objects.select_related('receta', 'bien').all()
    else:
        recetas_qs = Receta.objects.select_related('categoria').filter(usuario=usuario)
        productos_qs = Producto.objects.select_related('categoria').filter(usuario=usuario)
        relaciones_qs = ProductoReceta.objects.select_related('receta', 'producto').filter(receta__usuario=usuario)
        bienes_qs = Bien.objects.filter(usuario=usuario)
        relaciones_bienes_qs = BienReceta.objects.select_related('receta', 'bien').filter(receta__usuario=usuario)

    if request.method == 'POST':
        accion = request.POST.get('accion', '')

        # ------------------------------------------------------------------
        # EXPORTAR PLANTILLA
        # ------------------------------------------------------------------
        if accion == 'exportar_plantilla':
            wb = _build_workbook_recetas(
                productos=list(productos_qs.order_by('nombre')),
                bienes=list(bienes_qs.order_by('nombre')),
                configuracion=config,
                solo_plantilla=True,
            )
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="plantilla_recetas.xlsx"'
            wb.save(response)
            return response

        # ------------------------------------------------------------------
        # EXPORTAR RECETAS
        # ------------------------------------------------------------------
        if accion == 'exportar_recetas':
            recetas = recetas_qs.order_by('nombre')
            relaciones = relaciones_qs.order_by('receta__nombre', 'producto__nombre')
            wb = _build_workbook_recetas(
                productos=list(productos_qs.order_by('nombre')),
                recetas=list(recetas),
                relaciones=list(relaciones),
                bienes=list(bienes_qs.order_by('nombre')),
                relaciones_bienes=list(relaciones_bienes_qs.order_by('receta__nombre', 'bien__nombre')),
                configuracion=config,
            )
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="recetas.xlsx"'
            wb.save(response)
            return response

        # ------------------------------------------------------------------
        # PREVISUALIZAR IMPORTACIÓN
        # ------------------------------------------------------------------
        if accion == 'previsualizar':
            archivo = request.FILES.get('archivo_excel')
            if not archivo:
                messages.error(request, 'Debe seleccionar un archivo Excel.')
                return redirect('importar_exportar_recetas')

            if not archivo.name.lower().endswith('.xlsx'):
                messages.error(request, 'Solo se aceptan archivos .xlsx')
                return redirect('importar_exportar_recetas')

            _clear_preview_recetas(request, remove_file=True)
            temp_path = _store_preview_upload(archivo)
            preview = _analizar_archivo_recetas(temp_path, request)
            request.session[PREVIEW_RECETAS_SESSION_KEY] = _serialize_preview_for_session(
                preview,
                temp_path,
                archivo.name,
            )
            request.session.modified = True

            if preview['errors']:
                messages.error(request, 'Se detectaron errores. Revise la previsualización antes de confirmar.')
            elif preview['config_plan'] or preview['recipe_plans'] or preview['bien_plans'] or preview['relation_plans'] or preview['bien_relation_plans']:
                messages.success(request, 'Análisis completado. Puede revisar el preview antes de confirmar.')
            else:
                messages.info(request, 'No se encontraron filas con datos para analizar.')

            return redirect('importar_exportar_recetas')

        # ------------------------------------------------------------------
        # CONFIRMAR IMPORTACIÓN
        # ------------------------------------------------------------------
        if accion == 'confirmar_importacion':
            preview = request.session.get(PREVIEW_RECETAS_SESSION_KEY)
            token = request.POST.get('preview_token', '')
            if not preview or preview.get('token') != token:
                messages.error(request, 'La previsualización ya no es válida. Analice el archivo nuevamente.')
                _clear_preview_recetas(request, remove_file=True)
                return redirect('importar_exportar_recetas')

            success, analysis = _confirmar_importacion_recetas(preview, request)
            if not success:
                request.session[PREVIEW_RECETAS_SESSION_KEY] = _serialize_preview_for_session(
                    analysis,
                    preview['temp_path'],
                    preview.get('original_filename', 'archivo.xlsx'),
                )
                request.session.modified = True
                messages.error(request, 'La confirmación falló durante la revalidación. Revise el preview actualizado.')
                return redirect('importar_exportar_recetas')

            _clear_preview_recetas(request, remove_file=True)
            messages.success(
                request,
                'Importación completada: '
                f"{analysis['summary']['configuraciones_actualizar']} configuración(es) actualizada(s), "
                f"{analysis['summary']['configuraciones_crear']} configuración(es) creada(s), "
                f"{analysis['summary']['recetas_crear']} receta(s) creada(s), "
                f"{analysis['summary']['recetas_actualizar']} receta(s) actualizada(s), "
                f"{analysis['summary']['bienes_crear']} bien(es) creado(s), "
                f"{analysis['summary']['bienes_actualizar']} bien(es) actualizado(s), "
                f"{analysis['summary']['relaciones_crear']} ingrediente(s) creado(s), "
                f"{analysis['summary']['relaciones_actualizar']} ingrediente(s) actualizado(s), "
                f"{analysis['summary']['relaciones_bienes_crear']} bien(es)-receta creado(s) y "
                f"{analysis['summary']['relaciones_bienes_actualizar']} bien(es)-receta actualizado(s)."
            )
            return redirect('importar_exportar_recetas')

    # GET
    total_recetas = recetas_qs.count()
    preview = request.session.get(PREVIEW_RECETAS_SESSION_KEY)
    context = {
        'title': 'Importar / Exportar Recetas',
        'available_apps': admin.site.get_app_list(request),
        'total_recetas': total_recetas,
        'total_productos': productos_qs.count(),
        'config': config,
        'preview': preview,
    }
    return render(request, 'admin/importar_exportar_recetas.html', context)

